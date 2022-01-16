import xlsxwriter, io, time
from django.http import HttpResponse, request, response, HttpResponseRedirect
from django.template import RequestContext, context
from django.shortcuts import redirect, render
from .forms import RCAForm
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib import colors
from openpyxl import load_workbook
from .models import rca

def OnlyRCA(request):
    page_title = 'RCA'
    print('Pocela RCA funkcija...')
    rca_count = rca.objects.all().count()
    if request.method == "POST":
        form = RCAForm(request.POST, request.FILES)
        if form.is_valid():

            output = io.BytesIO()
            newdoc = request.FILES['file']

            p_start = time.time()

            wb = load_workbook(newdoc)
            if 'Instructions_READ_FIRST' in wb.sheetnames:
                wb.remove(wb['Instructions_READ_FIRST'])
            print ('Obrisan je sheet "Instructions_READ_FIRST" ukoliko je postojao!')

            dfs = pd.read_excel(wb, sheet_name=None, index_col=[0], engine='openpyxl')
            writer = pd.ExcelWriter(output)
            for name, df in dfs.items():
                print ('Sheet: ' + '"' + name + '"' + ' završen!') 
                                          
                df.columns = df.columns.str.split('_', expand=True)
                new_data = df.stack(0)
                new_data1 = new_data['uvoz_sektora'] = new_data['ImportValue'].sum()
                new_data2 = new_data['izvoz_sektora'] = new_data['ExportValue'].sum()
                new_data3 = new_data.eval('RCA = (ExportValue / izvoz_sektora) / (ImportValue / uvoz_sektora)')

                #ovo izbacuje kolone, za ovo ce primati argument s weba sta da izbaci
                cleaning_data = new_data3.drop(['ImportQuantity' , 'ImportValue', 'ExportQuantity', 'ExportValue', 'uvoz_sektora', 'izvoz_sektora'], axis = 1)

                #ovo vraca tabelu u normalu
                backtonormal = cleaning_data.unstack(1).swaplevel(0,1, axis=1).sort_index(axis=1)

                #stiliziranje
                cr = sns.light_palette("green", as_cmap=True)
                cr.set_bad('white')

                def my_gradient(s, cmap):
                    return [f'background-color: {colors.rgb2hex(x)}'
                            for x in cmap(s.replace(np.inf, np.nan))]

                styler = backtonormal.style
                zeleno = styler.apply(
                    my_gradient,
                    cmap=cr, 
                    subset=backtonormal.columns.get_loc_level('RCA', level=1)[0],
                    axis=0 
                                    )
                styler.to_excel(writer, sheet_name=name, float_format = "%0.2f")

            form.save()
            p_end = time.time()
            the_end = p_end - p_start
            print ('*' * 60)
            print('Funkcija je trajala ' + (str(the_end)[:-10]) + ' sekundi')
            print ('*' * 60)
            
            writer.save()
            output.seek(0)
            filename = f"RCA_calculated_{newdoc.name}"
            response = HttpResponse(
                output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=%s' % filename
            return response
    else:
        form = RCAForm()
    return render(request, 'only_rca.html', { 'form': form, 'page_title':page_title, 'rca_count':rca_count })

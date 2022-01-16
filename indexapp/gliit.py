import xlsxwriter, io, time
from django.http import HttpResponse, request, response, HttpResponseRedirect
from django.template import RequestContext, context
from django.shortcuts import redirect, render
from .forms import GLIITForm
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib import colors
from openpyxl import load_workbook
from .models import gliit

def OnlyGLIIT(request):
    page_title = 'GLIIT'
    print('Počela je GLIIT funkcija...')
    gliit_count = gliit.objects.all().count()
    if request.method == "POST":
        form = GLIITForm(request.POST, request.FILES)
        if form.is_valid():
            output = io.BytesIO()
            newdoc = request.FILES['file']

            p_start = time.time()

            wb = load_workbook(newdoc)
            if 'Instructions_READ_FIRST' in wb.sheetnames:
                wb.remove(wb['Instructions_READ_FIRST'])
            print ('Obrisan je sheet "Instructions_READ_FIRST" ukoliko je postojao!')

            dfs = pd.read_excel(wb, sheet_name=None, index_col=[0], engine='openpyxl')
            writer = pd.ExcelWriter(output)
            for name, df in dfs.items():
                print ('Sheet: ' + '"' + name + '"' + ' završen!')
                
                df.columns = df.columns.str.split('_', expand=True)
                new_data = df.stack(0)
                new_data1 = new_data.eval('abs_razlika = ExportValue - ImportValue').abs()
                new_data2 = new_data1.eval('abs_zbir = ExportValue + ImportValue').abs()
                new_data3 = new_data2.eval('GLIIT = (1 - (abs_razlika / abs_zbir)) * 100')
            
                #ovo izbacuje kolone, za ovo ce primati argument s weba sta da izbaci
                cleaning_data = new_data3.drop(['ImportQuantity' , 'ImportValue', 'ExportQuantity', 'ExportValue', 'abs_razlika','abs_zbir'], axis = 1)

                #ovo vraca tabelu u normalu
                backtonormal = cleaning_data.unstack(1).swaplevel(0,1, axis=1).sort_index(axis=1)
        
                #stiliziranje
                cc = sns.light_palette("red", as_cmap=True)
                cc.set_bad('white')

                def my_gradient(s, cmap):
                    my_norm=plt.Normalize(0, 100)
                    return [f'background-color: {colors.rgb2hex(x)}'  for x in cmap(my_norm(s.replace(np.inf, np.nan)) ) ]

                styler = backtonormal.style
                crveno = styler.apply(
                    my_gradient,
                    cmap=cc, 
                    subset=backtonormal.columns.get_loc_level('GLIIT', level=1)[0],
                    axis=0 
                                    )

                styler.to_excel(writer, sheet_name=name, float_format = "%0.2f")


            form.save()
            p_end = time.time()
            the_end = p_end - p_start
            print ('*' * 60)
            print('Funkcija je trajala ' + (str(the_end)[:-10]) + ' sekundi')
            print ('*' * 60)

            writer.save()
            output.seek(0)
            filename = f"GLIIT_calculated_{newdoc.name}"
            response = HttpResponse(
                output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=%s' % filename
            return response
    else:
        form = GLIITForm()
    return render(request, 'only_gliit.html', {'form': form, 'page_title':page_title, 'gliit_count':gliit_count })

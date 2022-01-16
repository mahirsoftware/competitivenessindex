import xlsxwriter, io, time 
from django.http import HttpResponse, request, response, HttpResponseRedirect
from django.template import RequestContext, context
from django.shortcuts import redirect, render
from .forms import TFCCForm
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib import colors
from openpyxl import load_workbook
from .models import tfcc

def OnlyTFCC(request):
    page_title = 'TFCC'
    print('Počela je TFCC funkcija...')
    tfcc_count = tfcc.objects.all().count()
    if request.method == "POST":
        form = TFCCForm(request.POST, request.FILES)
        if form.is_valid():
            p_start = time.time()

            output = io.BytesIO()
            newdoc = request.FILES['file']

            wb = load_workbook(newdoc)
            if 'Instructions_READ_FIRST' in wb.sheetnames:
                wb.remove(wb['Instructions_READ_FIRST'])
            print ('Obrisan je sheet "Instructions_READ_FIRST" ukoliko je postojao!')

            dfs = pd.read_excel(wb, sheet_name=None, index_col=[0], engine='openpyxl')
            writer = pd.ExcelWriter(output)
            for name, df in dfs.items():
                print ('Sheet: ' + '"' + name + '"' + ' završen!')
                
                df.columns = df.columns.str.split('_', expand=True)
                new_data = df.stack(0)
                new_data1 = new_data.eval('tb = ExportValue - ImportValue')
                new_data2 = new_data1.eval('uvX = ExportValue / ExportQuantity')
                new_data3 = new_data2.eval('uvM = ImportValue / ImportQuantity')
                new_data4 = new_data3.eval('uvd = uvX - uvM')
                new_data10 = new_data4
                uslovi = [
                    (new_data10['tb'] < 0) & (new_data10['ExportValue'] == 0), #samo uvozna TFCC, K1
                    (new_data10['tb'] < 0) & (new_data10['uvd'] < 0), # K2 
                    (new_data10['tb'] < 0) & (new_data10['uvd'] > 0), # K3       
                    (new_data10['tb'] == 0) & (new_data10['uvd'] < 0), # K4
                    (new_data10['tb'] < 0) & (new_data10['uvd'] == 0), #K5
                    (new_data10['tb'] == 0) & (new_data10['uvd'] == 0), #K6
                    (new_data10['tb'] > 0) & (new_data10['uvd'] == 0), #K7
                    (new_data10['tb'] == 0) & (new_data10['uvd'] > 0), # K8
                    (new_data10['tb'] > 0) & (new_data10['uvd'] < 0), # K9
                    (new_data10['tb'] > 0) & (new_data10['uvd'] > 0), # K10
                    (new_data10['tb'] > 0) & (new_data10['ImportValue'] == 0), # samo izvozna TFCC, K11
                ]
                rezultati = ['K1', 'K2', 'K3', 'K4', 'K5', 'K6', 'K7', 'K8', 'K9', 'K10', 'K11']
                new_data4['TFCC'] = np.select(uslovi, rezultati)
                #ovo izbacuje kolone, za ovo ce primati argument s weba sta da izbaci
                cleaning_data = new_data4.drop(['ImportQuantity' , 'ImportValue', 'ExportQuantity', 'ExportValue', 'tb', 'uvd', 'uvX', 'uvM'], axis = 1)

                #ovo vraca tabelu u normalu
                data_output = cleaning_data.unstack(1).swaplevel(0,1, axis=1).sort_index(axis=1)

                def highlight_TFCC(x):
                    color_data = pd.Series('background-color: white', index=x.index)
                    color_data[x.str.contains("K1", na=False)] = 'background-color: #e4f0f6'
                    color_data[x.str.contains("K2", na=False)] = 'background-color: #bcd9ea'
                    color_data[x.str.contains("K3", na=False)] = 'background-color: #8bbdd9'
                    color_data[x.str.contains("K4", na=False)] = 'background-color: #5ba4cf'
                    color_data[x.str.contains("K5", na=False)] = 'background-color: #298fca'
                    color_data[x.str.contains("K6", na=False)] = 'background-color: #0079bf'
                    color_data[x.str.contains("K7", na=False)] = 'background-color: #056cb8'
                    color_data[x.str.contains("K8", na=False)] = 'background-color: #026aa7'
                    color_data[x.str.contains("K9", na=False)] = 'background-color: #055a8c'
                    color_data[x.str.contains("K10", na=False)] = 'background-color: #094c72'
                    color_data[x.str.contains("K11", na=False)] = 'background-color: #0c3953'
                    return color_data

                idx = pd.IndexSlice

                styler = data_output.style.apply(highlight_TFCC, subset=idx[:, idx[:, 'TFCC']])
                
                styler.to_excel(writer, sheet_name=name)

            form.save()
            p_end = time.time()
            the_end = p_end - p_start
            print ('*' * 60)
            print('Funkcija je trajala ' + (str(the_end)[:-10]) + ' sekundi')
            print ('*' * 60)

            writer.save()
            output.seek(0)
            filename = f"TFCC_calculated_{newdoc.name}"
            response = HttpResponse(
                output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=%s' % filename
            return response
    else:
        form = TFCCForm()
    return render(request, 'only_tfcc.html', {'form': form, 'page_title':page_title, 'tfcc_count':tfcc_count })
